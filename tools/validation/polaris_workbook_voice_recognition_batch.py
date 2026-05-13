#!/usr/bin/env python3
from pathlib import Path
import sys

if __package__ in {None, ""}:
    ROOT = Path(__file__).resolve().parents[2]
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))

import argparse
import json
import re
import time
from collections import Counter
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

from tools.core.polaris_runtime import current_session_dir, workspace_root
from tools.device.polaris_network_orchestrator import hotspot_set, hotspot_status
from tools.probe.polaris_phrase_probe import run_probe


WORKBOOK_PREFIX = "20241128105427_"
FILTERED_JSON = "cache/runtime/workbook_20241128105427_voice_recognition_filtered.json"
REVIEW_JSON = "cache/runtime/workbook_20241128105427_voice_recognition_review.json"
SUMMARY_FILE = "voice_recognition_batch_summary.json"
SEND_VOICE = "\u53d1\u9001\u8bed\u97f3"
QUOTE_PAIRS = {
    '"': '"',
    "\u201c": "\u201d",
    "\u300c": "\u300d",
    "\u300e": "\u300f",
}
FALLBACK_TOKEN_RE = re.compile(r"Wakeup#talk#([^#]+)#.*?Asr#talk#([^#]+)#", re.IGNORECASE | re.DOTALL)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d%H%M%S")


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def workspace_path(*parts: str) -> Path:
    return workspace_root().joinpath(*parts)


def numeric_suffix(workbook_id: str) -> int:
    return int(str(workbook_id).rsplit("_", 1)[1])


def find_workbook() -> Path:
    matches = sorted(workspace_path("doc", "requirements").glob(f"{WORKBOOK_PREFIX}*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"missing workbook with prefix {WORKBOOK_PREFIX}")
    return matches[0]


def load_json(path: Path) -> object:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def normalized_steps(raw: object) -> str:
    text = str(raw or "")
    return text.replace("_x000D_", "\n").replace("\r\n", "\n").replace("\r", "\n")


def extract_spoken_phrase(steps: str) -> Optional[str]:
    steps = normalized_steps(steps)
    send_idx = steps.find(SEND_VOICE)
    search_start = send_idx if send_idx >= 0 else 0
    for pos in range(search_start, len(steps)):
        ch = steps[pos]
        close = QUOTE_PAIRS.get(ch)
        if not close:
            continue
        end = steps.find(close, pos + 1)
        if end > pos:
            return steps[pos + 1 : end].strip()
    token_match = FALLBACK_TOKEN_RE.search(steps)
    if token_match:
        wake = token_match.group(1).strip()
        asr = token_match.group(2).strip()
        if wake and asr:
            return f"{wake}\uff0c{asr}"
    return None


def load_workbook_rows(workbook_path: Path) -> Dict[str, dict]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: Dict[str, dict] = {}
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not header_seen:
            header_seen = True
            continue
        workbook_id = str(row[0] or "").strip()
        if not workbook_id:
            continue
        rows[workbook_id] = {
            "workbook_id": workbook_id,
            "category4": str(row[4] or ""),
            "type": str(row[5] or ""),
            "name": str(row[6] or ""),
            "steps": normalized_steps(row[9]),
            "expected_result": str(row[10] or ""),
        }
    wb.close()
    return rows


def pilot_summary_paths(session_dir: Path) -> Dict[str, Path]:
    base = session_dir / "artifacts" / "probe" / "phrase"
    return {
        "美的空调_T6挂机_716": base / "20260423154840629_phrase_probe_pilot_716_volume_percent" / "probe_summary.json",
        "美的空调_T6挂机_750": base / "20260423154908831_phrase_probe_pilot_750_multimedia" / "probe_summary.json",
        "美的空调_T6挂机_756": base / "20260423154934954_phrase_probe_pilot_756_weather" / "probe_summary.json",
    }


def expectation_for_case(workbook_id: str) -> dict:
    suffix = numeric_suffix(workbook_id)
    if 716 <= suffix <= 749:
        return {
            "family": "volume_percent_negative",
            "expected_keyword": "yin liang bai fen bi fan ji",
            "expected_tts_callback_id": 350,
        }
    if 750 <= suffix <= 756:
        return {
            "family": "online_skill_negative",
            "expected_keyword": "zai xian ji neng fan ji",
            "expected_tts_callback_id": 339,
        }
    raise ValueError(f"unsupported workbook_id for expectation: {workbook_id}")


def evaluate_step(
    case_item: dict,
    spoken: str,
    step_payload: dict,
    probe_summary_path: Path,
    *,
    reused_pilot: bool,
) -> dict:
    metrics = step_payload["metrics"]
    expectation = expectation_for_case(case_item["workbook_id"])

    checks = [
        {"name": "playback_returncode", "actual": step_payload["playback"]["returncode"], "expected": 0, "passed": step_payload["playback"]["returncode"] == 0},
        {"name": "cp_wake_count", "actual": metrics["cp_wake_count"], "expected": ">=1", "passed": metrics["cp_wake_count"] >= 1},
        {"name": "ap_wake_count", "actual": metrics["ap_wake_count"], "expected": ">=1", "passed": metrics["ap_wake_count"] >= 1},
        {"name": "cp_command_count", "actual": metrics["cp_command_count"], "expected": ">=1", "passed": metrics["cp_command_count"] >= 1},
        {"name": "ap_asr_count", "actual": metrics["ap_asr_count"], "expected": ">=1", "passed": metrics["ap_asr_count"] >= 1},
        {"name": "wb_asr_count", "actual": metrics["wb_asr_count"], "expected": ">=1", "passed": metrics["wb_asr_count"] >= 1},
        {
            "name": "recognized_command_keywords",
            "actual": metrics["recognized_command_keywords"],
            "expected": f"contains {expectation['expected_keyword']}",
            "passed": expectation["expected_keyword"] in set(metrics["recognized_command_keywords"]),
        },
        {
            "name": "wb_tts_callback_ids",
            "actual": metrics["wb_tts_callback_ids"],
            "expected": f"contains {expectation['expected_tts_callback_id']}",
            "passed": expectation["expected_tts_callback_id"] in set(metrics["wb_tts_callback_ids"]),
        },
        {"name": "boot_marker_count", "actual": metrics["boot_marker_count"], "expected": 0, "passed": metrics["boot_marker_count"] == 0},
        {"name": "crash_marker_count", "actual": metrics["crash_marker_count"], "expected": 0, "passed": metrics["crash_marker_count"] == 0},
    ]

    failed = [item for item in checks if not item["passed"]]
    verdict = "PASS" if not failed else "FAIL"
    return {
        "action": "executed",
        "workbook_id": case_item["workbook_id"],
        "case_id": case_item.get("case_id", ""),
        "mode": case_item.get("mode", ""),
        "category4": case_item.get("category4", ""),
        "type": case_item.get("type", ""),
        "name": case_item.get("name", ""),
        "spoken": spoken,
        "family": expectation["family"],
        "verdict": verdict,
        "checks": checks,
        "failed_checks": failed,
        "probe_summary_path": str(probe_summary_path),
        "probe_execution_dir": str(probe_summary_path.parent),
        "probe_step_id": step_payload["step_id"],
        "probe_step_dir": str(probe_summary_path.parent / step_payload["step_id"]),
        "reused_pilot": reused_pilot,
        "metrics": {
            "cp_wake_count": metrics["cp_wake_count"],
            "ap_wake_count": metrics["ap_wake_count"],
            "cp_command_count": metrics["cp_command_count"],
            "ap_asr_count": metrics["ap_asr_count"],
            "wb_asr_count": metrics["wb_asr_count"],
            "recognized_command_keywords": metrics["recognized_command_keywords"],
            "wb_tts_callback_ids": metrics["wb_tts_callback_ids"],
            "boot_marker_count": metrics["boot_marker_count"],
            "crash_marker_count": metrics["crash_marker_count"],
        },
    }


def make_error_result(case_item: dict, spoken: Optional[str], error: str) -> dict:
    return {
        "action": "executed",
        "workbook_id": case_item["workbook_id"],
        "case_id": case_item.get("case_id", ""),
        "mode": case_item.get("mode", ""),
        "category4": case_item.get("category4", ""),
        "type": case_item.get("type", ""),
        "name": case_item.get("name", ""),
        "spoken": spoken,
        "verdict": "ERROR",
        "error": error,
    }


def wait_for_hotspot_state(
    *,
    expect_operational_state: Optional[str] = None,
    expect_client_count: Optional[int] = None,
    expect_client_mac: Optional[str] = None,
    timeout_s: float = 90.0,
    interval_s: float = 2.0,
) -> Tuple[dict, bool]:
    deadline = time.time() + timeout_s
    last = hotspot_status()
    while time.time() <= deadline:
        last = hotspot_status()
        ok = True
        if expect_operational_state is not None and str(last.get("operational_state")) != expect_operational_state:
            ok = False
        if expect_client_count is not None and int(last.get("client_count", -1)) != expect_client_count:
            ok = False
        if expect_client_mac is not None:
            clients = [str(item.get("mac_address", "")).lower() for item in last.get("clients", [])]
            if expect_client_mac.lower() not in clients:
                ok = False
        if ok:
            return last, True
        time.sleep(interval_s)
    return last, False


def build_validation_dir(session_dir: Path) -> Path:
    root = session_dir / "artifacts" / "validation" / "workbook_20241128105427_voice_recognition_batch"
    path = root / f"{now_stamp()}_batch_rerun"
    path.mkdir(parents=True, exist_ok=False)
    return path


def classify_all_cases(filtered_cases: List[dict], executed_map: Dict[str, dict]) -> List[dict]:
    all_cases: List[dict] = []
    for item in filtered_cases:
        workbook_id = item["workbook_id"]
        if workbook_id in executed_map:
            all_cases.append(executed_map[workbook_id])
            continue
        classification = item.get("classification")
        if classification == "auto_executable_now":
            all_cases.append(
                {
                    "action": "skip_already_covered",
                    "workbook_id": workbook_id,
                    "case_id": item.get("case_id", ""),
                    "mode": item.get("mode", ""),
                    "category4": item.get("category4", ""),
                    "type": item.get("type", ""),
                    "name": item.get("name", ""),
                    "previous_result": item.get("result", ""),
                }
            )
        elif classification == "skip":
            all_cases.append(
                {
                    "action": "skip_not_executable",
                    "workbook_id": workbook_id,
                    "case_id": item.get("case_id", ""),
                    "mode": item.get("mode", ""),
                    "category4": item.get("category4", ""),
                    "type": item.get("type", ""),
                    "name": item.get("name", ""),
                    "skip_reason": item.get("reason", ""),
                }
            )
        else:
            all_cases.append(
                {
                    "action": "unhandled_missing",
                    "workbook_id": workbook_id,
                    "case_id": item.get("case_id", ""),
                    "mode": item.get("mode", ""),
                    "category4": item.get("category4", ""),
                    "type": item.get("type", ""),
                    "name": item.get("name", ""),
                }
            )
    return all_cases


def run_batch(observe_ms: int) -> Path:
    session_dir = current_session_dir()
    workbook_path = find_workbook()
    filtered_cases = list(load_json(workspace_path(FILTERED_JSON)))
    review_payload = dict(load_json(workspace_path(REVIEW_JSON)))
    workbook_rows = load_workbook_rows(workbook_path)
    validation_dir = build_validation_dir(session_dir)

    summary = {
        "generated_at": now_iso(),
        "source_workbook": str(workbook_path),
        "source_filtered_json": str(workspace_path(FILTERED_JSON)),
        "source_review_json": str(workspace_path(REVIEW_JSON)),
        "session_dir": str(session_dir),
        "validation_dir": str(validation_dir),
        "filter_rule": "voice_interaction/speech_recognition/online_or_offline/exclude_stability_and_pressure",
        "handled_total": len(filtered_cases),
        "review_counts": review_payload.get("classification_counts", {}),
        "hotspot_before": None,
        "hotspot_off_result": None,
        "hotspot_offline_state": None,
        "hotspot_offline_ready": None,
        "hotspot_restore_result": None,
        "hotspot_after_restore": None,
        "hotspot_restore_ready": None,
        "pending_batch_probe_summary": None,
        "executed_results": [],
        "all_cases": [],
        "action_counts": {},
        "executed_counts": {},
        "executed_pass_family_counts": {},
    }

    env_payload = {}
    env_path = workspace_path("config", "polaris_env.json")
    if env_path.exists():
        env_payload = json.loads(env_path.read_text(encoding="utf-8"))
    dut_mac = str(env_payload.get("current_deviceinfo", {}).get("mac", "")).lower()

    executed_results: List[dict] = []
    executed_map: Dict[str, dict] = {}
    batch_inputs: List[dict] = []
    pilot_paths = pilot_summary_paths(session_dir)

    try:
        candidate_cases = [
            item
            for item in filtered_cases
            if item.get("classification") == "missing" and 716 <= numeric_suffix(item["workbook_id"]) <= 756
        ]
        candidate_cases.sort(key=lambda item: numeric_suffix(item["workbook_id"]))

        for case_item in candidate_cases:
            workbook_id = case_item["workbook_id"]
            wb_row = workbook_rows.get(workbook_id)
            if not wb_row:
                result = make_error_result(case_item, None, "workbook row not found")
                executed_results.append(result)
                executed_map[workbook_id] = result
                continue
            spoken = extract_spoken_phrase(wb_row["steps"])
            if not spoken:
                result = make_error_result(case_item, None, "cannot extract spoken phrase from workbook steps")
                executed_results.append(result)
                executed_map[workbook_id] = result
                continue
            pilot_path = pilot_paths.get(workbook_id)
            if pilot_path and pilot_path.exists():
                pilot_summary = json.loads(pilot_path.read_text(encoding="utf-8"))
                step_payload = pilot_summary["steps"][0]
                result = evaluate_step(case_item, step_payload["text"], step_payload, pilot_path, reused_pilot=True)
                executed_results.append(result)
                executed_map[workbook_id] = result
                continue
            batch_inputs.append({"case_item": case_item, "spoken": spoken})

        write_json(validation_dir / "batch_case_plan.json", {"cases": batch_inputs})

        summary["hotspot_before"] = hotspot_status()
        if str(summary["hotspot_before"].get("operational_state")) == "On":
            summary["hotspot_off_result"] = hotspot_set(False)
            offline_state, offline_ready = wait_for_hotspot_state(
                expect_operational_state="Off",
                expect_client_count=0,
                timeout_s=45.0,
            )
            summary["hotspot_offline_state"] = offline_state
            summary["hotspot_offline_ready"] = offline_ready
        else:
            summary["hotspot_off_result"] = {"skipped": True, "reason": "hotspot already off"}
            summary["hotspot_offline_state"] = hotspot_status()
            summary["hotspot_offline_ready"] = True

        if batch_inputs:
            label = "wb20241128105427_voice_recognition_remaining"
            probe_summary_path = run_probe(
                texts=[item["spoken"] for item in batch_inputs],
                device_key="",
                observe_ms=observe_ms,
                label=label,
            )
            summary["pending_batch_probe_summary"] = str(probe_summary_path)
            probe_summary = json.loads(probe_summary_path.read_text(encoding="utf-8"))
            for batch_input, step_payload in zip(batch_inputs, probe_summary["steps"]):
                case_item = batch_input["case_item"]
                result = evaluate_step(case_item, batch_input["spoken"], step_payload, probe_summary_path, reused_pilot=False)
                executed_results.append(result)
                executed_map[case_item["workbook_id"]] = result

        all_cases = classify_all_cases(filtered_cases, executed_map)
        summary["executed_results"] = sorted(executed_results, key=lambda item: numeric_suffix(item["workbook_id"]))
        summary["all_cases"] = all_cases
        summary["action_counts"] = dict(Counter(item["action"] for item in all_cases))
        summary["executed_counts"] = dict(Counter(item.get("verdict", "") for item in summary["executed_results"]))
        summary["executed_pass_family_counts"] = dict(
            Counter(item["family"] for item in summary["executed_results"] if item.get("verdict") == "PASS" and item.get("family"))
        )
        write_json(validation_dir / SUMMARY_FILE, summary)
    finally:
        try:
            before = summary.get("hotspot_before") or {}
            restore_needed = str(before.get("operational_state")) == "On"
            if restore_needed:
                summary["hotspot_restore_result"] = hotspot_set(True)
                after_restore, restore_ready = wait_for_hotspot_state(
                    expect_operational_state="On",
                    expect_client_mac=dut_mac if dut_mac else None,
                    timeout_s=120.0,
                )
                summary["hotspot_after_restore"] = after_restore
                summary["hotspot_restore_ready"] = restore_ready
            else:
                summary["hotspot_restore_result"] = {"skipped": True, "reason": "hotspot was not on before batch"}
                summary["hotspot_after_restore"] = hotspot_status()
                summary["hotspot_restore_ready"] = True
        finally:
            write_json(validation_dir / SUMMARY_FILE, summary)

    return validation_dir / SUMMARY_FILE


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run workbook 20241128105427 voice recognition batch validation")
    parser.add_argument("--observe-ms", type=int, default=15000)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = run_batch(observe_ms=args.observe_ms)
    print(result)


if __name__ == "__main__":
    main()
