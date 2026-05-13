# -*- coding: utf-8 -*-
from pathlib import Path
import sys

if __package__ in {None, ""}:
    ROOT = Path(__file__).resolve().parents[2]
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))

import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from tools.core.polaris_runtime import find_artifact_dirs, workspace_root


STEP_RE = re.compile(
    r"(Wakeup|Asr|UnAsr|online_Asr|online_UnAsr|Action|Check)#(talk|shell|sleep|regex)#([^#]*)#"
)

MODE_OFFLINE = "离线"
MODE_SEMI_OFFLINE = "在离线"
MODE_ONLINE = "在线"
MODE_MANUAL = "手动执行"
MODE_AUTO = "自动执行"

MODEL_CODE_TO_NAME = {
    "CA3X": "CA3X系列空调",
}

COLMO_MODELS = {
    "厨房空调",
    "EVO挂机",
    "EVO柜机",
}

SINGLE_MIC_MODELS = {
    "单麦钰行挂机",
}

SUPPORTED_DOC_CASES: Dict[str, dict] = {
    "美的空调_28": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 14000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_tones": [3],
        "min_wb_playback_end": 1,
        "confidence": "high",
        "notes": "文档离线 one-shot，唤醒后 10ms 接打开空调。",
    },
    "美的空调_29": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 14000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_tones": [3],
        "min_wb_playback_end": 1,
        "confidence": "high",
        "notes": "文档离线 one-shot，唤醒后 1s 接打开空调。",
    },
    "美的空调_30": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 18000,
        "min_cp_wake": 1,
        "min_unique_command_keywords": 2,
        "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji"],
        "required_tones": [4],
        "min_wb_playback_end": 3,
        "confidence": "medium",
        "notes": "文档离线 one-shot，关闭空调后再打开空调；判定重点是两条命令是否都被识别并播报。",
    },
    "美的空调_31": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 12000,
        "min_cp_wake": 4,
        "min_wb_wake": 4,
        "min_wb_playback_end": 1,
        "confidence": "medium",
        "notes": "文档离线唤醒提示语打断，重点看连续唤醒链路是否不断。",
    },
    "美的空调_32": {
        "runner_kind": "offline_interrupt_voice",
        "observe_after_ms": 14000,
        "min_cp_wake": 2,
        "min_wb_wake": 2,
        "min_wb_playback_end": 2,
        "require_wake_during_playback": True,
        "confidence": "medium",
        "notes": "文档离线识别播报语打断，重点验证命令播报开始后第二次唤醒是否能打断当前播报。",
    },
    "美的空调_42": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 16000,
        "min_cp_wake": 6,
        "min_wb_wake": 6,
        "min_wb_playback_end": 6,
        "confidence": "high",
        "notes": "文档离线连续 6 次唤醒。",
    },
    "美的空调_43": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 24000,
        "min_cp_wake": 3,
        "min_wb_wake": 3,
        "min_wb_playback_end": 3,
        "confidence": "high",
        "notes": "文档离线连续 3 次唤醒。",
    },
    "美的空调_44": {
        "runner_kind": "online_stress_case",
        "scenario": "wake_only",
        "stress_cycles": 1000,
        "stress_seed": 44,
        "observe_after_ms": 10000,
        "min_cp_wake": 1000,
        "min_ap_wake": 1000,
        "min_wb_online_wake": 1000,
        "min_ap_instruction_broadcast": 950,
        "max_boot_markers": 0,
        "max_crash_markers": 0,
        "confidence": "medium",
        "notes": "在线连续 1000 轮纯唤醒压测：至少保持 1000 次 CP/AP/WB 在线唤醒命中，且云端唤醒播报链路基本不掉次，不出现异常重启或崩溃标记。",
    },
    "美的空调_45": {
        "runner_kind": "online_stress_case",
        "scenario": "wake_command_interrupt",
        "stress_cycles": 1000,
        "stress_seed": 45,
        "observe_after_ms": 10000,
        "min_cp_wake": 1000,
        "min_cp_command": 1000,
        "min_ap_wake": 1000,
        "min_wb_online_wake": 1000,
        "min_unique_command_keywords": 2,
        "required_command_keywords": ["kong tiao kai ji", "kong tiao guan ji"],
        "min_ap_instruction_broadcast": 900,
        "min_interrupt_reset_count": 100,
        "max_boot_markers": 0,
        "max_crash_markers": 0,
        "confidence": "medium",
        "notes": "在线连续 1000 轮唤醒+识别压测，并把部分下一轮唤醒压到上一轮识别播报窗口内；要求命令识别与云端播报持续可用，不出现异常重启或崩溃标记。",
    },
    "美的空调_48": {
        "runner_kind": "offline_voice",
        "observe_after_ms": 14000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_tones": [3],
        "min_wb_playback_end": 1,
        "confidence": "high",
        "notes": "文档离线识别测试示例，先回归当前已知命令词打开空调。",
    },
    "美的空调_51": {
        "runner_kind": "serial_only",
        "observe_after_ms": 8000,
        "min_wb_tts_callback": 1,
        "min_wb_playback_start": 1,
        "min_wb_playback_end": 1,
        "confidence": "medium",
        "notes": "文档离线 TTS 播报语示例，当前自动验证示例命令 listen player play 310 的链路是否真的可播报。",
    },
    "美的空调_1": {
        "runner_kind": "power_broadcast_case",
        "network_state": "offline",
        "disconnect_wait_s": 20.0,
        "reconnect_wait_s": 70.0,
        "power_off_wait_s": 2.0,
        "power_observe_s": 25.0,
        "min_wb_playback_end": 2,
        "required_tones": [102, 406],
        "forbidden_tones": [290],
        "confidence": "medium",
        "notes": "热点离线后执行 WB01 硬重启，校验上电欢迎链路是否走“欢迎使用美的语音空调 + 未联网提示”播报。",
    },
    "美的空调_5": {
        "runner_kind": "power_broadcast_case",
        "network_state": "online",
        "power_off_wait_s": 2.0,
        "power_observe_s": 25.0,
        "min_wb_playback_end": 2,
        "required_tones": [102, 290],
        "forbidden_tones": [406],
        "confidence": "medium",
        "notes": "联网状态下执行 WB01 硬重启，校验上电欢迎链路是否走“欢迎使用美的语音空调 + 主人请吩咐”播报。",
    },
    "美的空调_20": {
        "runner_kind": "network_reconnect_voice_case",
        "disconnect_wait_s": 25.0,
        "reconnect_wait_s": 70.0,
        "offline_observe_after_ms": 12000,
        "online_observe_after_ms": 12000,
        "confidence": "medium",
        "notes": "热点断开后，在线技能“现在几点了”应不再进入在线 ASR/云端 TTS，但离线“打开空调”仍可控制；热点恢复后，在线技能与空调控制都应恢复正常。",
    },
    "美的空调_21": {
        "runner_kind": "online_empty_nlu_case",
        "observe_after_ms": 8000,
        "min_cp_wake": 4,
        "min_ap_wake": 4,
        "min_wb_online_wake": 4,
        "required_online_asr_texts": ["do"],
        "min_ap_cloud_tts_play": 4,
        "max_unique_command_keywords": 0,
        "confidence": "medium",
        "notes": "在线状态下连续 4 轮唤醒后说“度”，应稳定形成在线 ASR 文本 do，并进入云端兜底 TTS，而不是命中本地空调命令。",
    },
    "美的空调_52": {
        "runner_kind": "app_mic_case",
        "scenario": "mic_off_reminder_window",
        "cloud_apply_wait_s": 6.0,
        "observe_after_ms": 6000,
        "min_cp_wake": 4,
        "max_ap_wake": 0,
        "max_wb_wake": 0,
        "max_ap_asr": 0,
        "max_wb_asr": 0,
        "required_tones": [417],
        "min_wb_playback_start": 1,
        "max_wb_playback_start": 1,
        "min_wb_playback_end": 1,
        "max_wb_playback_end": 1,
        "confidence": "medium",
        "notes": "APP 关闭语音后，单次唤醒不应触发 AP/WB 播报；10s 内连续 3 次唤醒应只播报一次 tone 417，11s 后再唤醒不应重复提醒。",
    },
    "美的空调_53": {
        "runner_kind": "app_mic_case",
        "scenario": "mic_off_persist_after_power_cycle_online",
        "cloud_apply_wait_s": 6.0,
        "observe_after_ms": 6000,
        "min_cp_wake": 1,
        "max_ap_wake": 0,
        "max_wb_wake": 0,
        "max_ap_asr": 0,
        "max_wb_asr": 0,
        "max_wb_playback_start": 0,
        "max_wb_playback_end": 0,
        "confidence": "medium",
        "notes": "APP 关闭语音后执行 WB01 掉电上电，重启后单次唤醒应仍只留 CP wake，不应恢复用户可感知播报。",
    },
    "美的空调_54": {
        "runner_kind": "app_mic_case",
        "scenario": "mic_on_online_command",
        "cloud_apply_wait_s": 6.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_unique_command_keywords": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "min_ap_cloud_tts_play": 1,
        "confidence": "medium",
        "notes": "APP 打开语音后，在线唤醒+打开空调应恢复正常交互，至少出现一次有效命令识别和云端 TTS 播放。",
    },
    "美的空调_55": {
        "runner_kind": "app_mic_case",
        "scenario": "mic_on_offline_interaction",
        "cloud_apply_wait_s": 6.0,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 6000,
        "min_cp_wake": 4,
        "min_ap_wake": 1,
        "min_unique_command_keywords": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "required_tones": [3],
        "forbidden_tones": [417],
        "min_wb_playback_end": 1,
        "confidence": "medium",
        "notes": "APP 打开语音后断网，离线唤醒+打开空调仍应可用，且连续唤醒不应出现“语音已关闭”提醒 tone 417。",
    },
    "美的空调_56": {
        "runner_kind": "app_mic_case",
        "scenario": "mic_off_persist_after_power_cycle_offline",
        "cloud_apply_wait_s": 6.0,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 6000,
        "min_cp_wake": 1,
        "max_ap_wake": 0,
        "max_wb_wake": 0,
        "max_ap_asr": 0,
        "max_wb_asr": 0,
        "max_wb_playback_start": 0,
        "max_wb_playback_end": 0,
        "confidence": "medium",
        "notes": "APP 关闭语音后先断网再执行 WB01 掉电上电，离线重启后单次唤醒仍不应恢复任何 AP/WB 播报。",
    },
    "美的空调_58": {
        "runner_kind": "app_proactive_mic_case",
        "cloud_apply_wait_s": 6.0,
        "observe_per_phase_s": 6.0,
        "confidence": "medium",
        "notes": "APP 语音关闭时 4 组主动交互都不应播报；重新打开语音后同 4 组主动交互都应恢复播报。",
    },
    "美的空调_123": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 30,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "required_tones": [287],
        "min_wb_playback_end": 2,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 30s，再断网；离线仅唤醒后应播报超时退出提示音。",
    },
    "美的空调_124": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 30,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "required_tones": [3, 287],
        "min_wb_playback_end": 3,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 30s，再断网；离线唤醒+打开空调后应先播报命令结果，再播报超时退出提示音。",
    },
    "美的空调_125": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 20,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "required_tones": [287],
        "min_wb_playback_end": 2,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 20s，再断网；离线仅唤醒后应播报超时退出提示音。",
    },
    "美的空调_126": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 20,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "required_tones": [3, 287],
        "min_wb_playback_end": 3,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 20s，再断网；离线唤醒+打开空调后应先播报命令结果，再播报超时退出提示音。",
    },
    "美的空调_127": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 15,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "required_tones": [287],
        "min_wb_playback_end": 2,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 15s，再断网；离线仅唤醒后应播报超时退出提示音。",
    },
    "美的空调_128": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 15,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "required_tones": [3, 287],
        "min_wb_playback_end": 3,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 15s，再断网；离线唤醒+打开空调后应先播报命令结果，再播报超时退出提示音。",
    },
    "美的空调_129": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 10,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "required_tones": [287],
        "min_wb_playback_end": 2,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 10s，再断网；离线仅唤醒后应播报超时退出提示音。",
    },
    "美的空调_130": {
        "runner_kind": "app_offline_timeout_case",
        "timeout_seconds": 10,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "required_tones": [3, 287],
        "min_wb_playback_end": 3,
        "confidence": "medium",
        "notes": "APP 先打开自然对话并设超时 10s，再断网；离线唤醒+打开空调后应先播报命令结果，再播报超时退出提示音。",
    },
    "美的空调_298": {
        "runner_kind": "dialog_phase_case",
        "scenario": "dialog_timeout",
        "dialog_mode": "half",
        "noise_mode": "silent",
        "confidence": "medium",
        "notes": "离线半双工：关闭自然对话后，验证空调开机时有超时播报、关机时无超时播报。",
    },
    "美的空调_299": {
        "runner_kind": "dialog_phase_case",
        "scenario": "dialog_timeout",
        "dialog_mode": "half",
        "noise_mode": "many",
        "confidence": "medium",
        "notes": "离线半双工：关闭自然对话后，验证 3 次及以上非空调人声时的兜底+超时逻辑。",
    },
    "美的空调_300": {
        "runner_kind": "dialog_phase_case",
        "scenario": "dialog_timeout",
        "dialog_mode": "half",
        "noise_mode": "few",
        "confidence": "medium",
        "notes": "离线半双工：关闭自然对话后，验证少量非空调人声时直接超时的逻辑。",
    },
    "美的空调_301": {
        "runner_kind": "dialog_phase_case",
        "scenario": "half_duplex_first_command_only",
        "confidence": "medium",
        "notes": "离线半双工：同一会话多条指令仅第一条应生效。",
    },
    "美的空调_302": {
        "runner_kind": "dialog_phase_case",
        "scenario": "dialog_timeout",
        "dialog_mode": "full",
        "noise_mode": "silent",
        "confidence": "medium",
        "notes": "离线全双工：打开自然对话后，验证空调开机时有超时播报、关机时无超时播报。",
    },
    "美的空调_303": {
        "runner_kind": "dialog_phase_case",
        "scenario": "dialog_timeout",
        "dialog_mode": "full",
        "noise_mode": "many",
        "confidence": "medium",
        "notes": "离线全双工：打开自然对话后，验证 3 次及以上非空调人声时的兜底+超时逻辑。",
    },
    "美的空调_304": {
        "runner_kind": "dialog_phase_case",
        "scenario": "dialog_timeout",
        "dialog_mode": "full",
        "noise_mode": "few",
        "confidence": "medium",
        "notes": "离线全双工：打开自然对话后，验证少量非空调人声时直接超时的逻辑。",
    },
    "美的空调_311": {
        "runner_kind": "dialog_phase_case",
        "scenario": "switch_effect",
        "switch_from": "full",
        "switch_to": "half",
        "confidence": "medium",
        "notes": "离线全双工切半双工：关闭自然对话后，后续同一句里的打开空调不应再识别。",
    },
    "美的空调_312": {
        "runner_kind": "dialog_phase_case",
        "scenario": "switch_effect",
        "switch_from": "half",
        "switch_to": "full",
        "confidence": "medium",
        "notes": "离线半双工切全双工：打开自然对话后，后续同一句里的打开空调不应在当句立即生效。",
    },
    "美的空调_613": {
        "runner_kind": "dialog_phase_case",
        "scenario": "stress_interaction",
        "dialog_mode": "full",
        "stress_cycles": 8,
        "stress_seed": 613,
        "confidence": "medium",
        "notes": "未联网+自然对话打开的离线语音交互压测：执行多轮随机间隔的唤醒+空调命令，校验不重启、不死机、链路持续可用。",
    },
    "美的空调_614": {
        "runner_kind": "dialog_phase_case",
        "scenario": "stress_interaction",
        "dialog_mode": "half",
        "stress_cycles": 8,
        "stress_seed": 614,
        "confidence": "medium",
        "notes": "未联网+自然对话关闭的离线语音交互压测：执行多轮随机间隔的唤醒+空调命令，校验不重启、不死机、链路持续可用。",
    },
    "美的空调_685": {
        "runner_kind": "app_wakeup_word_persist_case",
        "target_wakeup_word": "客厅空调",
        "probe_text": "客厅空调",
        "recovery_wakeup_word": "小美小美",
        "cloud_apply_wait_s": 12.0,
        "observe_after_ms": 10000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_wb_playback_end": 1,
        "confidence": "medium",
        "notes": "APP 切换唤醒词后执行 WB01 掉电上电，重启后切换后的唤醒词仍应可用。",
    },
    "美的空调_686": {
        "runner_kind": "app_threshold_persist_case",
        "probe_text": "小美小美",
        "expected_wakeup_keyword": "xiao mei xiao mei",
        "pre_threshold_high": 100,
        "pre_threshold_low": 0,
        "expected_wakeup_threshold": -73,
        "expected_recognition_threshold": -308,
        "double_wake_gap_ms": 1500,
        "recovery_threshold": 50,
        "cloud_apply_wait_s": 12.0,
        "observe_after_ms": 10000,
        "min_cp_wake": 2,
        "min_ap_wake": 2,
        "confidence": "medium",
        "notes": "APP 先把唤醒阈值切到最高再切到最低并掉电重启，重启后最低阈值应保留并继续生效。",
    },
    "美的空调_687": {
        "runner_kind": "app_accent_persist_case",
        "full_duplex_enable": False,
        "timeout_seconds": 15,
        "cloud_apply_wait_s": 6.0,
        "cloud_recovery_wait_s": 6.0,
        "observe_after_ms": 12000,
        "power_off_wait_s": 2.0,
        "power_observe_s": 25.0,
        "mixed_res_enable": 0,
        "restore_accent_id": "cantonese",
        "accent_plan": [
            {"accent_id": "cantonese", "label": "粤语"},
            {"accent_id": "henanhua", "label": "河南话"},
            {"accent_id": "shanghaihua", "label": "上海话"},
            {"accent_id": "shandonghua", "label": "山东话"},
            {"accent_id": "minnanhua", "label": "闽南话"},
        ],
        "confidence": "medium",
        "notes": "APP 逐一切换方言后执行 WB01 掉电上电；在当前缺少方言音频的口径下，以重启后的 cloud.order.config.query.reply 仍保留目标 accentId/enableAccent，且“小美小美打开空调”继续表现为方言开启后的 one-shot 降级作为 PASS 证据。",
    },
    "美的空调_585": {
        "runner_kind": "network_disconnect_case",
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 70.0,
        "confidence": "medium",
        "notes": "热点断网后应先在 AP 日志看到 AI disconnected，再在 WB01 日志看到 class ai state 4，对应文档里的 ai,4 断云返回码。",
    },
    "美的空调_706": {
        "runner_kind": "wake_info_upload_case",
        "wake_text": "小美小美",
        "observe_after_ms": 10000,
        "confidence": "medium",
        "notes": "在线唤醒后应在 AP 日志同时看到本地 algo info 与发送到云端的 device.report.wakeInfo，关键字段保持一致。",
    },
    "美的空调_707": {
        "runner_kind": "algo_version_upload_case",
        "power_off_wait_s": 2.0,
        "power_observe_s": 25.0,
        "confidence": "medium",
        "notes": "在线硬重启后应在 AP 日志看到上传的 algo_version / esrVersion，并与本地 version 命令输出保持一致。",
    },
    "美的空调_709": {
        "runner_kind": "cloud_log_upload_probe_case",
        "log_status": 1,
        "log_level": 7,
        "expected_device_loglev": 4,
        "clear_commands": [],
        "probe_rounds": 2,
        "observe_after_ms": 12000,
        "restore_default_after": True,
        "confidence": "medium",
        "notes": "在线将云端日志上传等级切到 debug 后，本地应看到 `set device loglev 4 by cloud_change`，并能继续完成在线语音交互；最终仍需云端日志回捞才能判 PASS。",
    },
    "美的空调_710": {
        "runner_kind": "cloud_log_upload_probe_case",
        "log_status": 1,
        "log_level": 6,
        "expected_device_loglev": 3,
        "clear_commands": [],
        "probe_rounds": 2,
        "observe_after_ms": 12000,
        "restore_default_after": True,
        "confidence": "medium",
        "notes": "在线将云端日志上传等级切到 info 后，本地应看到 `set device loglev 3 by cloud_change`，并能继续完成在线语音交互；最终仍需云端日志回捞才能判 PASS。",
    },
    "美的空调_711": {
        "runner_kind": "cloud_log_upload_probe_case",
        "log_status": 1,
        "log_level": 4,
        "expected_device_loglev": 2,
        "clear_commands": [],
        "probe_rounds": 2,
        "observe_after_ms": 12000,
        "restore_default_after": True,
        "confidence": "medium",
        "notes": "在线将云端日志上传等级切到 warning 后，本地应看到 `set device loglev 2 by cloud_change`，并能继续完成在线语音交互；最终仍需云端日志回捞才能判 PASS。",
    },
    "美的空调_712": {
        "runner_kind": "cloud_log_upload_probe_case",
        "log_status": 1,
        "log_level": 3,
        "expected_device_loglev": 1,
        "clear_commands": [],
        "probe_rounds": 2,
        "observe_after_ms": 12000,
        "restore_default_after": True,
        "confidence": "medium",
        "notes": "在线将云端日志上传等级切到 error 后，本地应看到 `set device loglev 1 by cloud_change`，并能继续完成在线语音交互；最终仍需云端日志回捞才能判 PASS。",
    },
    "美的空调_713": {
        "runner_kind": "cloud_log_upload_probe_case",
        "log_status": 0,
        "log_level": 7,
        "expected_device_loglev": 1,
        "clear_commands": [],
        "probe_rounds": 2,
        "observe_after_ms": 12000,
        "restore_default_after": True,
        "confidence": "medium",
        "notes": "在线恢复默认日志上传配置后，本地应看到 `set device loglev 1 by cloud_change`，并能继续完成在线语音交互；最终仍需云端日志回捞才能判 PASS。",
    },
    "美的空调_714": {
        "runner_kind": "wakeup_audio_upload_probe_case",
        "enable_upload": True,
        "probe_rounds": 1,
        "probe_gap_ms": 1800,
        "tail_silence_ms": 300000,
        "observe_after_ms": 15000,
        "confidence": "medium",
        "notes": "在线开启唤醒音频上传后，本地应至少看到一次 `wakeup_upload` 成功响应与 `isUploadingFile=1`；最终仍需云端下载音频并校验格式/内容才能判 PASS。",
    },
    "美的空调_704": {
        "runner_kind": "app_accent_case",
        "scenario": "accent_blocks_oneshot",
        "full_duplex_enable": False,
        "timeout_seconds": 15,
        "cloud_apply_wait_s": 6.0,
        "cloud_recovery_wait_s": 6.0,
        "observe_after_ms": 12000,
        "mixed_res_enable": 0,
        "restore_accent_id": "cantonese",
        "accent_plan": [
            {"accent_id": "cantonese", "label": "粤语"},
            {"accent_id": "henanhua", "label": "河南话"},
            {"accent_id": "shanghaihua", "label": "上海话"},
            {"accent_id": "shandonghua", "label": "山东话"},
            {"accent_id": "minnanhua", "label": "闽南话"},
        ],
        "confidence": "medium",
        "notes": "打开方言后回归在线 one-shot：对粤语/河南话/上海话/山东话/闽南话逐一切换，再验证“小美小美打开空调”不再保留完整命令词在线 ASR。",
    },
    "美的空调_705": {
        "runner_kind": "app_accent_case",
        "scenario": "accent_off_supports_oneshot",
        "full_duplex_enable": False,
        "timeout_seconds": 15,
        "cloud_apply_wait_s": 6.0,
        "cloud_recovery_wait_s": 6.0,
        "observe_after_ms": 12000,
        "mixed_res_enable": 0,
        "pre_enable_accent_id": "cantonese",
        "restore_accent_id": "cantonese",
        "confidence": "medium",
        "notes": "先切到方言再关闭方言恢复普通话，验证“小美小美打开空调”重新支持在线 one-shot，在线 ASR 只保留“打开空调”。",
    },
}

SUPPORTED_DOC_CASES.update(
    {
        "美的空调_33": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "observe_after_ms": 12000,
            "min_cp_wake": 2,
            "min_ap_wake": 2,
            "min_ap_cloud_tts_play": 1,
            "required_online_asr_texts": ["合肥今天的天气"],
            "min_interrupt_reset_count": 1,
            "confidence": "medium",
            "notes": "在线天气播报期间再次唤醒，应打断当前在线播报并切回新的唤醒会话。",
        },
        "美的空调_34": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "observe_after_ms": 12000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_ap_cloud_tts_play": 2,
            "required_online_asr_texts": ["合肥今天的天气", "今天的股票情况"],
            "min_interrupt_reset_count": 1,
            "confidence": "medium",
            "notes": "在线全双工下，天气播报期间继续说股票查询，应打断前一条在线播报并切换到后一条请求。",
        },
        "美的空调_22": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 half-duplex 下，唤醒后 10ms 内说“打开空调”应形成稳定识别与云端播报。",
        },
        "美的空调_23": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 half-duplex 下，唤醒后 1s 说“打开空调”应形成稳定识别与云端播报。",
        },
        "美的空调_24": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 half-duplex 下仅唤醒并等待超时，再说命令词不应被识别，且应只播报一次超时退出提示。",
        },
        "美的空调_25": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 full-duplex 下，唤醒后 10ms 内说“打开空调”应形成稳定识别与云端播报。",
        },
        "美的空调_26": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 full-duplex 下，唤醒后 1s 说“打开空调”应形成稳定识别与云端播报。",
        },
        "美的空调_27": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 2,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji"],
            "min_ap_cloud_tts_play": 2,
            "confidence": "medium",
            "notes": "在线 full-duplex 下，单次唤醒后先说“关闭空调”再说“打开空调”应都被识别并播报。",
        },
        "美的空调_50": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_ap_cloud_tts_play": 1,
            "required_online_asr_texts": ["帮我定个明天早上七点的闹钟"],
            "confidence": "medium",
            "notes": "在线闹钟技能应至少完成一次稳定识别，并返回云端播报结果。",
        },
        "美的空调_114": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 30,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 3,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji", "zhi leng mo shi"],
            "min_ap_cloud_tts_play": 3,
            "confidence": "medium",
            "notes": "APP 打开自然对话 30s 后，单次唤醒下连续三条在线空调指令都应可用。",
        },
        "美的空调_115": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 30,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 打开自然对话 30s 后，仅唤醒并等待 30s 应只播报一次超时退出提示。",
        },
        "美的空调_116": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 30,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 2,
            "confidence": "medium",
            "notes": "APP 打开自然对话 30s 后，先说“打开空调”再等待 30s，应先有命令播报，再有超时退出提示。",
        },
        "美的空调_117": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 20,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 打开自然对话 20s 后，仅唤醒并等待 20s 应只播报一次超时退出提示。",
        },
        "美的空调_118": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 20,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 2,
            "confidence": "medium",
            "notes": "APP 打开自然对话 20s 后，先说“打开空调”再等待 20s，应先有命令播报，再有超时退出提示。",
        },
        "美的空调_119": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 打开自然对话 15s 后，仅唤醒并等待 15s 应只播报一次超时退出提示。",
        },
        "美的空调_120": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 2,
            "confidence": "medium",
            "notes": "APP 打开自然对话 15s 后，先说“打开空调”再等待 15s，应先有命令播报，再有超时退出提示。",
        },
        "美的空调_121": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 10,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 打开自然对话 10s 后，仅唤醒并等待 10s 应只播报一次超时退出提示。",
        },
        "美的空调_122": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 10,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 2,
            "confidence": "medium",
            "notes": "APP 打开自然对话 10s 后，先说“打开空调”再等待 10s，应先有命令播报，再有超时退出提示。",
        },
        "美的空调_139": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "max_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao guan ji"],
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 关闭自然对话后，首条“关闭空调”应识别播报，15s 后再说“打开空调”不应继续识别。",
        },
        "美的空调_131": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 10,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 打开自然对话 10s 后，等待超时再说“打开空调”不应继续识别。",
        },
        "美的空调_132": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 3,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji", "zhi leng mo shi"],
            "min_ap_cloud_tts_play": 3,
            "confidence": "medium",
            "notes": "APP 打开自然对话 15s 后，连续三条在线空调指令都应可识别并播报。",
        },
        "美的空调_677": {
            "runner_kind": "app_mic_case",
            "scenario": "mic_on_persist_after_power_cycle_online",
            "cloud_apply_wait_s": 6.0,
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao kai ji"],
            "min_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 先关再开语音后掉电重启，重启后在线唤醒+命令应恢复正常交互。",
        },
        "美的空调_678": {
            "runner_kind": "app_mic_case",
            "scenario": "mic_off_toggle_persist_after_power_cycle_online",
            "cloud_apply_wait_s": 6.0,
            "observe_after_ms": 6000,
            "min_cp_wake": 1,
            "max_ap_wake": 0,
            "max_wb_wake": 0,
            "max_ap_asr": 0,
            "max_wb_asr": 0,
            "max_wb_playback_start": 0,
            "max_wb_playback_end": 0,
            "confidence": "medium",
            "notes": "APP 先开再关语音后掉电重启，重启后单次唤醒仍应保持静默。",
        },
        "美的空调_681": {
            "runner_kind": "app_dialog_persist_case",
            "precondition_method": "cloud",
            "dialog_mode": "half",
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "max_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao guan ji"],
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 关闭自然对话后掉电重启，重启后应保持一次唤醒一次识别。",
        },
        "美的空调_682": {
            "runner_kind": "app_dialog_persist_case",
            "precondition_method": "cloud",
            "dialog_mode": "full",
            "timeout_seconds": 15,
            "observe_after_ms": 12000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 3,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji", "zhi leng mo shi"],
            "min_ap_cloud_tts_play": 3,
            "confidence": "medium",
            "notes": "APP 打开自然对话后掉电重启，重启后应保持单次唤醒多轮识别。",
        },
        "美的空调_683": {
            "runner_kind": "app_dialog_persist_case",
            "precondition_method": "voice",
            "dialog_mode": "half",
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "max_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao guan ji"],
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "语音关闭自然对话后掉电重启，重启后应保持一次唤醒一次识别。",
        },
        "美的空调_684": {
            "runner_kind": "app_dialog_persist_case",
            "precondition_method": "voice",
            "dialog_mode": "full",
            "timeout_seconds": 15,
            "observe_after_ms": 12000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 3,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji", "zhi leng mo shi"],
            "min_ap_cloud_tts_play": 3,
            "confidence": "medium",
            "notes": "语音打开自然对话后掉电重启，重启后应保持单次唤醒多轮识别。",
        },
    }
)

SUPPORTED_DOC_CASES.update(
    {
        "美的空调_46": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，打开空调",
            "prepare_command_keywords": ["kong tiao kai ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 half-duplex 唤醒后等待超时，应出现一次退出提示播报。",
        },
        "美的空调_47": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，打开空调",
            "prepare_command_keywords": ["kong tiao kai ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 3,
            "min_ap_wake": 3,
            "max_unique_command_keywords": 0,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "在线 half-duplex 连续三次唤醒后，最后一轮应仍能稳定进入并按时退出会话。",
        },
        "美的空调_113": {
            "runner_kind": "app_dialog_announce_case",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "expected_wb_fullduplex": 2,
            "confidence": "medium",
            "notes": "APP 打开自然对话后，应立即在 AP/WB 日志看到 full-duplex 配置生效与对应播报开始/结束链路。",
        },
        "美的空调_133": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 20,
            "observe_after_ms": 10000,
            "min_cp_wake": 2,
            "min_ap_wake": 2,
            "max_unique_command_keywords": 0,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "APP 将自然对话最终配置为 20s 后，超时后再说“打开空调”不应继续识别。",
        },
        "美的空调_134": {
            "runner_kind": "app_dialog_announce_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_enable_full_duplex": True,
            "expected_wb_fullduplex": 1,
            "confidence": "medium",
            "notes": "APP 关闭自然对话后，应立即在 AP/WB 日志看到 half-duplex 配置生效与对应播报开始链路。",
        },
        "美的空调_135": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": True,
            "timeout_seconds": 10,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "max_ap_cloud_tts_play": 0,
            "confidence": "medium",
            "notes": "空调处于关机态时，full-duplex 10s 仅唤醒等待不应再播报超时提示。",
        },
        "美的空调_137": {
            "runner_kind": "app_dialog_config_case",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_command_text": "小美小美，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "max_ap_cloud_tts_play": 0,
            "confidence": "medium",
            "notes": "空调处于关机态时，half-duplex 15s 仅唤醒等待不应再播报超时提示。",
        },
        "美的空调_140": {
            "runner_kind": "app_dialog_config_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "full_duplex_enable": True,
            "timeout_seconds": 30,
            "observe_after_ms": 12000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 3,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji", "zhi leng mo shi"],
            "min_ap_cloud_tts_play": 3,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo + full-duplex 30s 后，三条在线空调指令都应可识别并播报。",
        },
        "美的空调_157": {
            "runner_kind": "app_offline_timeout_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "timeout_seconds": 10,
            "disconnect_wait_s": 15.0,
            "reconnect_wait_s": 60.0,
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_wb_wake": 1,
            "required_tones": [287],
            "min_wb_playback_end": 2,
            "max_unique_command_keywords": 0,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo 并断网后，10s 超时退出后再说“打开空调”不应继续识别。",
        },
        "美的空调_158": {
            "runner_kind": "app_dialog_config_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "full_duplex_enable": True,
            "timeout_seconds": 15,
            "observe_after_ms": 12000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 3,
            "required_command_keywords": ["kong tiao guan ji", "kong tiao kai ji", "zhi leng mo shi"],
            "min_ap_cloud_tts_play": 3,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo + full-duplex 15s 后，三条在线空调指令都应可识别并播报。",
        },
        "美的空调_159": {
            "runner_kind": "app_dialog_config_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "full_duplex_enable": True,
            "timeout_seconds": 20,
            "observe_after_ms": 10000,
            "min_cp_wake": 2,
            "min_ap_wake": 2,
            "max_unique_command_keywords": 0,
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo 并把自然对话设为 20s 后，超时后再说“打开空调”不应继续识别。",
        },
        "美的空调_160": {
            "runner_kind": "app_dialog_config_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "observe_after_ms": 10000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "min_unique_command_keywords": 1,
            "max_unique_command_keywords": 1,
            "required_command_keywords": ["kong tiao guan ji"],
            "min_ap_cloud_tts_play": 1,
            "max_ap_cloud_tts_play": 1,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo 并关闭自然对话后，首条“关闭空调”应识别，15s 后再说“打开空调”不应继续识别。",
        },
        "美的空调_161": {
            "runner_kind": "app_dialog_config_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "full_duplex_enable": True,
            "timeout_seconds": 10,
            "prepare_command_text": "hicolmo，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "max_ap_cloud_tts_play": 0,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo 且空调处于关机态时，full-duplex 10s 仅唤醒等待不应再播报超时提示。",
        },
        "美的空调_163": {
            "runner_kind": "app_dialog_config_case",
            "target_wakeup_word": "hicolmo",
            "recovery_wakeup_word": "小美小美",
            "full_duplex_enable": False,
            "timeout_seconds": 15,
            "prepare_command_text": "hicolmo，关闭空调",
            "prepare_command_keywords": ["kong tiao guan ji"],
            "observe_after_ms": 8000,
            "min_cp_wake": 1,
            "min_ap_wake": 1,
            "max_unique_command_keywords": 0,
            "max_ap_cloud_tts_play": 0,
            "confidence": "medium",
            "notes": "切换唤醒词 hicolmo 且空调处于关机态时，half-duplex 15s 仅唤醒等待不应再播报超时提示。",
        },
    }
)

for case_id, timeout_seconds in [
    ("美的空调_141", 30),
    ("美的空调_143", 20),
    ("美的空调_145", 15),
    ("美的空调_147", 10),
]:
    SUPPORTED_DOC_CASES[case_id] = {
        "runner_kind": "app_dialog_config_case",
        "target_wakeup_word": "hicolmo",
        "recovery_wakeup_word": "小美小美",
        "full_duplex_enable": True,
        "timeout_seconds": timeout_seconds,
        "observe_after_ms": 10000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_ap_cloud_tts_play": 1,
        "max_ap_cloud_tts_play": 1,
        "confidence": "medium",
        "notes": f"切换唤醒词 hicolmo + full-duplex {timeout_seconds}s 后，仅唤醒并等待应只播报一次超时退出提示。",
    }

for case_id, timeout_seconds in [
    ("美的空调_142", 30),
    ("美的空调_144", 20),
    ("美的空调_146", 15),
    ("美的空调_148", 10),
]:
    SUPPORTED_DOC_CASES[case_id] = {
        "runner_kind": "app_dialog_config_case",
        "target_wakeup_word": "hicolmo",
        "recovery_wakeup_word": "小美小美",
        "full_duplex_enable": True,
        "timeout_seconds": timeout_seconds,
        "observe_after_ms": 10000,
        "min_cp_wake": 1,
        "min_ap_wake": 1,
        "min_unique_command_keywords": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "min_ap_cloud_tts_play": 2,
        "confidence": "medium",
        "notes": f"切换唤醒词 hicolmo + full-duplex {timeout_seconds}s 后，先说“打开空调”再等待，应先有命令播报，再有超时退出提示。",
    }

for case_id, timeout_seconds in [
    ("美的空调_149", 30),
    ("美的空调_151", 20),
    ("美的空调_153", 15),
    ("美的空调_155", 10),
]:
    SUPPORTED_DOC_CASES[case_id] = {
        "runner_kind": "app_offline_timeout_case",
        "target_wakeup_word": "hicolmo",
        "recovery_wakeup_word": "小美小美",
        "timeout_seconds": timeout_seconds,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "required_tones": [287],
        "min_wb_playback_end": 2,
        "confidence": "medium",
        "notes": f"切换唤醒词 hicolmo 并断网后，离线自然对话 {timeout_seconds}s 仅唤醒等待应播报超时退出提示。",
    }

for case_id, timeout_seconds in [
    ("美的空调_150", 30),
    ("美的空调_152", 20),
    ("美的空调_154", 15),
    ("美的空调_156", 10),
]:
    SUPPORTED_DOC_CASES[case_id] = {
        "runner_kind": "app_offline_timeout_case",
        "target_wakeup_word": "hicolmo",
        "recovery_wakeup_word": "小美小美",
        "timeout_seconds": timeout_seconds,
        "disconnect_wait_s": 15.0,
        "reconnect_wait_s": 60.0,
        "observe_after_ms": 8000,
        "min_cp_wake": 1,
        "min_wb_wake": 1,
        "min_ap_asr": 1,
        "min_wb_asr": 1,
        "required_command_keywords": ["kong tiao kai ji"],
        "required_tones": [3, 287],
        "min_wb_playback_end": 3,
        "confidence": "medium",
        "notes": f"切换唤醒词 hicolmo 并断网后，离线自然对话 {timeout_seconds}s 先说“打开空调”再等待，应先播报命令结果，再播报超时退出提示。",
    }

WAKE_WORD_CASE_MATRIX = {
    "美的空调_61": {"target_wakeup_word": "卧室空调", "probe_text": "卧室空调", "expect_wake": True},
    "美的空调_62": {"target_wakeup_word": "卧室空调", "probe_text": "小美小美", "expect_wake": False},
    "美的空调_63": {"target_wakeup_word": "卧室空调", "probe_text": "客厅空调", "expect_wake": False},
    "美的空调_64": {"target_wakeup_word": "卧室空调", "probe_text": "书房空调", "expect_wake": False},
    "美的空调_65": {"target_wakeup_word": "客厅空调", "probe_text": "客厅空调", "expect_wake": True},
    "美的空调_66": {"target_wakeup_word": "客厅空调", "probe_text": "小美小美", "expect_wake": False},
    "美的空调_67": {"target_wakeup_word": "客厅空调", "probe_text": "卧室空调", "expect_wake": False},
    "美的空调_68": {"target_wakeup_word": "客厅空调", "probe_text": "书房空调", "expect_wake": False},
    "美的空调_69": {"target_wakeup_word": "书房空调", "probe_text": "书房空调", "expect_wake": True},
    "美的空调_70": {"target_wakeup_word": "书房空调", "probe_text": "小美小美", "expect_wake": False},
    "美的空调_71": {"target_wakeup_word": "书房空调", "probe_text": "卧室空调", "expect_wake": False},
    "美的空调_72": {"target_wakeup_word": "书房空调", "probe_text": "客厅空调", "expect_wake": False},
    "美的空调_73": {"target_wakeup_word": "小美小美", "probe_text": "小美小美", "expect_wake": True},
    "美的空调_74": {"target_wakeup_word": "小美小美", "probe_text": "书房空调", "expect_wake": False},
    "美的空调_75": {"target_wakeup_word": "小美小美", "probe_text": "卧室空调", "expect_wake": False},
    "美的空调_76": {"target_wakeup_word": "小美小美", "probe_text": "客厅空调", "expect_wake": False},
}

for case_id, wake_rule in WAKE_WORD_CASE_MATRIX.items():
    target_wakeup_word = wake_rule["target_wakeup_word"]
    probe_text = wake_rule["probe_text"]
    expect_wake = bool(wake_rule["expect_wake"])
    notes = (
        f"APP 将唤醒词切换为 {target_wakeup_word} 后，语音 {probe_text} 应能正常唤醒并播报提示音。"
        if expect_wake
        else f"APP 将唤醒词切换为 {target_wakeup_word} 后，语音 {probe_text} 不应触发用户可感知的唤醒播报。"
    )
    rule = {
        "runner_kind": "app_wakeup_word_case",
        "target_wakeup_word": target_wakeup_word,
        "probe_text": probe_text,
        "recovery_wakeup_word": "小美小美",
        "cloud_apply_wait_s": 12.0,
        "observe_after_ms": 10000,
        "confidence": "medium",
        "notes": notes,
    }
    if expect_wake:
        rule.update(
            {
                "min_cp_wake": 1,
                "min_ap_wake": 1,
                "min_ap_cloud_tts_start": 1,
                "min_ap_cloud_tts_stop": 1,
            }
        )
    else:
        rule.update(
            {
                "max_ap_wake": 0,
                "max_wb_wake": 0,
                "max_ap_asr": 0,
                "max_wb_asr": 0,
                "max_wb_playback_start": 0,
                "max_wb_playback_end": 0,
            }
        )
    SUPPORTED_DOC_CASES[case_id] = rule

THRESHOLD_REQUEST_MAPPING = {
    0: 0,
    1: 25,
    2: 50,
    3: 75,
    4: 100,
}

THRESHOLD_CASE_GROUPS = [
    {
        "case_ids": ["美的空调_77", "美的空调_78", "美的空调_79", "美的空调_80", "美的空调_81"],
        "target_wakeup_word": "小美小美",
        "probe_text": "小美小美",
        "expected_keyword": "xiao mei xiao mei",
        "expected_thresholds": [-73, -82, -90, -100, -109],
    },
    {
        "case_ids": ["美的空调_87", "美的空调_88", "美的空调_89", "美的空调_90", "美的空调_91"],
        "target_wakeup_word": "hicolmo",
        "probe_text": "hicolmo",
        "expected_keyword": "hi colmo",
        "expected_thresholds": [-74, -83, -91, -101, -110],
    },
    {
        "case_ids": ["美的空调_92", "美的空调_93", "美的空调_94", "美的空调_95", "美的空调_96"],
        "target_wakeup_word": "你好科慕",
        "probe_text": "你好科慕",
        "expected_keyword": "ni hao ke mu",
        "expected_thresholds": [-138, -149, -160, -173, -187],
    },
    {
        "case_ids": ["美的空调_97", "美的空调_98", "美的空调_99", "美的空调_100", "美的空调_101"],
        "target_wakeup_word": "客厅空调",
        "probe_text": "客厅空调",
        "expected_keyword": "ke ting kong tiao",
        "expected_thresholds": [-153, -165, -177, -204, -235],
    },
    {
        "case_ids": ["美的空调_102", "美的空调_103", "美的空调_104", "美的空调_105", "美的空调_106"],
        "target_wakeup_word": "书房空调",
        "probe_text": "书房空调",
        "expected_keyword": "shu fang kong tiao",
        "expected_thresholds": [-153, -165, -177, -204, -235],
    },
    {
        "case_ids": ["美的空调_107", "美的空调_108", "美的空调_109", "美的空调_110", "美的空调_111"],
        "target_wakeup_word": "卧室空调",
        "probe_text": "卧室空调",
        "expected_keyword": "wo shi kong tiao",
        "expected_thresholds": [-147, -158, -170, -197, -226],
    },
]

for group in THRESHOLD_CASE_GROUPS:
    group_id = f"threshold_group_{group['expected_keyword'].replace(' ', '_')}"
    for index, case_id in enumerate(group["case_ids"]):
        threshold_request = THRESHOLD_REQUEST_MAPPING[index]
        expected_threshold = group["expected_thresholds"][index]
        SUPPORTED_DOC_CASES[case_id] = {
            "runner_kind": "app_threshold_case",
            "target_wakeup_word": group["target_wakeup_word"],
            "probe_text": group["probe_text"],
            "expected_wakeup_keyword": group["expected_keyword"],
            "threshold_request": threshold_request,
            "expected_wakeup_threshold": expected_threshold,
            "expected_recognition_threshold": -308,
            "threshold_group_id": group_id,
            "threshold_group_index": index,
            "threshold_group_case_ids": group["case_ids"],
            "threshold_group_reference_thresholds": group["expected_thresholds"],
            "double_wake_gap_ms": 1500,
            "recovery_wakeup_word": "小美小美",
            "recovery_threshold": 50,
            "cloud_apply_wait_s": 12.0,
            "observe_after_ms": 10000,
            "min_cp_wake": 2,
            "min_ap_wake": 2,
            "confidence": "medium",
            "notes": (
                f"APP 设置唤醒词为 {group['target_wakeup_word']}、阈值请求为 {threshold_request} 后，"
                f"语音 {group['probe_text']} 的首次阈值应为 {expected_threshold}，"
                "同一会话内二次唤醒应出现识别态阈值 -308。"
            ),
        }

UNSUPPORTED_REASON_OVERRIDES: Dict[str, str] = {
    "美的空调_1": "需要人工断电上电，当前持续日志会话下不自动执行电源循环。",
    "美的空调_3": "当前已支持自动断电上电，但该用例仍要求听辨上电播报、唤醒提示语和命令词提示语是否为默认发音人；现有日志只能确认默认配置与播报链路，缺少客观的发音人身份标记。",
    "美的空调_4": "当前已支持自动断电上电，但该用例仍要求听辨上电播报、唤醒提示语和命令词提示语是否为默认发音人；现有日志只能确认默认配置与播报链路，缺少客观的发音人身份标记。",
    "美的空调_6": "前置条件要求清除配网信息，当前未自动改写设备网络状态。",
    "美的空调_36": "该用例前置操作依赖“功能键/确认键”让设备进入配网等待态，当前未接入遥控器或面板按键自动化入口。",
    "美的空调_40": "该用例要求把设备录音线接到电脑并做 8 通道录音拆分，当前未接入对应的多通道采集链路与自动判音能力。",
    "美的空调_41": "该用例要求把设备录音线接到电脑并做 8 通道录音拆分，当前未接入对应的多通道采集链路与自动判音能力。",
    "美的空调_57": "该用例要求遍历上电播报、闹钟（默认/音乐）、主动交互、方言切换、联网播报等整组功能；当前缺少闹钟/UAT 环境切换的稳定自动编排，只能做局部验证，暂不标记为全自动。",
    "美的空调_679": "该用例要求通过遥控器关闭再打开语音后掉电，当前缺少遥控器/面板侧的自动化入口；不是电源控制能力不足。",
    "美的空调_680": "该用例要求通过遥控器打开再关闭语音后掉电，当前缺少遥控器/面板侧的自动化入口；不是电源控制能力不足。",
    "美的空调_688": "当前设备按本地文档提供的音色枚举发起音色切换时，云端接口返回 code 501/未找到对应的音色，无法稳定建立“切换音色后再掉电”的前置态。",
    "美的空调_708": "文档本身仍写“操作待定/跳过”，当前缺少可执行且可客观判定的异常注入步骤。",
    "美的空调_212": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_214": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_216": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_218": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_220": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_222": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_224": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_226": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_228": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
    "美的空调_230": "已支持通过 APP/cloud 脚本设置方言与唤醒词前置，但该用例还要求删设备后断网校验，当前未接入删设备/断网编排。",
}

SAFE_SHELL_COMMANDS = {
    "console 1",
    "flash.setloglev 4",
    "listen player play 310",
    "listen player play 76",
    "listen flash setloglev 4",
    "listen version",
    "listen flash show",
    "version",
    "deviceinfo",
    "flash.show",
    "player.setloglev 4",
    "mai.setloglev 4",
}

DEFERRED_STABILITY_RUNNER_KINDS = {
    "online_stress_case",
}

APP_KEYWORDS = ["手机app", "美居app", "美的美居", "app上", "在手机app"]
POWER_KEYWORDS = ["断电", "上电", "插上电源", "拔掉电源", "掉电"]
REMOTE_KEYWORDS = ["遥控器", "功能键", "确认键", "按键", "按钮"]
NETWORK_KEYWORDS = ["联网", "路由器", "wifi", "ai云", "iot", "热点", "post请求"]
MANUAL_KEYWORDS = ["听一下", "观察", "找实习生", "人工", "扫码", "查看"]
OTA_KEYWORDS = ["ota", "升级", "回滚"]


@dataclass
class StepToken:
    kind: str
    channel: str
    value: str


@dataclass
class DocCase:
    case_id: str
    level1: str
    level2: str
    level3: str
    level4: str
    case_type: str
    name: str
    priority: str
    precondition: str
    steps: str
    expected: str
    notes: str
    owner: str
    tokens: List[StepToken]

    def as_dict(self) -> dict:
        payload = asdict(self)
        payload["tokens"] = [asdict(token) for token in self.tokens]
        return payload


def default_doc_xlsx() -> Path:
    case_dir = workspace_root() / "doc" / "cases"
    candidates = sorted(case_dir.glob("*.xlsx"))
    if not candidates:
        legacy = sorted((workspace_root() / "doc").glob("*.xlsx"))
        if legacy:
            return legacy[0]
        raise FileNotFoundError(f"no doc-case xlsx found under {case_dir}")
    preferred_markers = ("??????", "???????", "????(?)")
    for marker in preferred_markers:
        for candidate in candidates:
            if marker in candidate.name:
                return candidate
    return candidates[0]


def normalize_device_model_name(text: str) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    upper = raw.upper()
    for code, model_name in MODEL_CODE_TO_NAME.items():
        if code in upper:
            return model_name
    return raw


def extract_model_code_from_ota_url(url: str) -> str:
    match = re.search(r"/Midea_OTA/([^/]+)/", str(url or ""), re.IGNORECASE)
    if not match:
        return ""
    return match.group(1).strip()


def infer_device_model(env: Optional[dict] = None) -> Tuple[str, str]:
    env = env or {}
    for key in ["current_device_model", "device_model", "current_model_label"]:
        model = normalize_device_model_name(str(env.get(key, "")))
        if model:
            return model, f"env.{key}"

    session_candidates = [
        str(env.get("latest_state_probe", "")).strip(),
        str(env.get("last_state_probe", "")).strip(),
    ]
    active_session = str(env.get("active_result_session", "")).strip()
    if active_session:
        session_candidates.append(active_session)

    checked_dirs = set()
    for candidate in session_candidates:
        if not candidate:
            continue
        path = Path(candidate)
        session_dir = path.parent if path.is_file() else path
        if not session_dir.exists():
            continue
        session_key = str(session_dir.resolve())
        if session_key in checked_dirs:
            continue
        checked_dirs.add(session_key)
        state_files = sorted(
            [json_path for probe_dir in find_artifact_dirs("state_probe", session_dir=session_dir) for json_path in probe_dir.glob("*.json")],
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        for state_file in state_files:
            try:
                payload = json.loads(state_file.read_text(encoding="utf-8"))
            except Exception:
                continue
            ota_url = (
                payload.get("wb01", {})
                .get("listen_flash_show", {})
                .get("ota_url", "")
            ) or (
                payload.get("wb", {})
                .get("listen_flash_show", {})
                .get("ota_url", "")
            )
            model = normalize_device_model_name(extract_model_code_from_ota_url(ota_url))
            if model:
                return model, f"{state_file.name}:wb.listen_flash_show.ota_url"
    return "", ""


def build_device_capability_tags(model_name: str) -> Dict[str, bool]:
    model = normalize_device_model_name(model_name)
    return {
        "is_colmo_model": model in COLMO_MODELS,
        "is_single_mic_model": model in SINGLE_MIC_MODELS,
    }


def classify_model_applicability(case: DocCase, env: Optional[dict] = None) -> Optional[dict]:
    model_name, model_source = infer_device_model(env)
    if not model_name:
        return None

    capability = build_device_capability_tags(model_name)
    notes = str(case.notes or "")
    normalized = notes.lower()
    source_desc = f"{model_name}（来源 {model_source}）" if model_source else model_name

    if "非colmo机型跳过此用例" in normalized and not capability["is_colmo_model"]:
        return {
            "status": "skip",
            "reason": f"用例备注标注“非colmo机型跳过此用例”；当前设备机型识别为 {source_desc}，不适用。",
            "runner_kind": "",
        }

    if "colmo机型跳过此用例" in normalized and capability["is_colmo_model"]:
        return {
            "status": "skip",
            "reason": f"用例备注标注“colmo机型跳过此用例”；当前设备机型识别为 {source_desc}，不适用。",
            "runner_kind": "",
        }

    if "单麦" in normalized and "其他机型没有此功能" in normalized and not capability["is_single_mic_model"]:
        return {
            "status": "skip",
            "reason": f"用例备注标注单麦专属功能（其他机型没有此功能）；当前设备机型识别为 {source_desc}，不适用。",
            "runner_kind": "",
        }

    case_rule = SUPPORTED_DOC_CASES.get(case.case_id, {})
    target_wakeup_word = str(case_rule.get("target_wakeup_word", ""))
    if (
        case_rule.get("runner_kind") == "app_wakeup_word_persist_case"
        and target_wakeup_word in {"卧室空调", "客厅空调", "书房空调"}
        and not capability["is_single_mic_model"]
    ):
        return {
            "status": "skip",
            "reason": (
                f"该掉电保持用例依赖房间名唤醒词 `{target_wakeup_word}`；同族用例备注已标注为单麦专属能力，"
                f"当前设备机型识别为 {source_desc}，不适用。"
            ),
            "runner_kind": "",
        }

    if (
        "colmo机型（厨房空调、evo柜机、evo挂机）使用此用例，其他机型跳过此用例" in normalized
        and model_name not in COLMO_MODELS
    ):
        return {
            "status": "skip",
            "reason": f"用例备注标注仅 colmo 机型（厨房空调 / EVO柜机 / EVO挂机）适用；当前设备机型识别为 {source_desc}，不适用。",
            "runner_kind": "",
        }

    if (
        "colmo机型（厨房空调、evo柜机、evo挂机）跳过此用例" in normalized
        and model_name in COLMO_MODELS
    ):
        return {
            "status": "skip",
            "reason": f"用例备注标注 colmo 机型（厨房空调 / EVO柜机 / EVO挂机）跳过；当前设备机型识别为 {source_desc}，不适用。",
            "runner_kind": "",
        }

    return None



def parse_tokens(text: str) -> List[StepToken]:
    return [StepToken(kind=m.group(1), channel=m.group(2), value=m.group(3)) for m in STEP_RE.finditer(text or "")]



def load_doc_cases(xlsx_path: Optional[Path] = None) -> List[DocCase]:
    xlsx_path = xlsx_path or default_doc_xlsx()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    cases: List[DocCase] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        case = DocCase(
            case_id="" if row[0] is None else str(row[0]).strip(),
            level1="" if row[1] is None else str(row[1]).strip(),
            level2="" if row[2] is None else str(row[2]).strip(),
            level3="" if row[3] is None else str(row[3]).strip(),
            level4="" if row[4] is None else str(row[4]).strip(),
            case_type="" if row[5] is None else str(row[5]).strip(),
            name="" if row[6] is None else str(row[6]).strip(),
            priority="" if row[7] is None else str(row[7]).strip(),
            precondition="" if row[8] is None else str(row[8]).strip(),
            steps="" if row[9] is None else str(row[9]).strip(),
            expected="" if row[10] is None else str(row[10]).strip(),
            notes="" if row[11] is None else str(row[11]).strip(),
            owner="" if row[12] is None else str(row[12]).strip(),
            tokens=parse_tokens("" if row[9] is None else str(row[9])),
        )
        cases.append(case)
    return cases



def load_doc_case(case_id: str, xlsx_path: Optional[Path] = None) -> DocCase:
    for case in load_doc_cases(xlsx_path=xlsx_path):
        if case.case_id == case_id:
            return case
    raise KeyError(f"doc case not found: {case_id}")



def case_text(case: DocCase) -> str:
    return "\n".join(
        [case.level1, case.level2, case.level3, case.level4, case.case_type, case.name, case.precondition, case.steps, case.expected, case.notes]
    )


def stability_execution_policy(env: Optional[dict] = None) -> dict:
    env = env or {}
    policy = dict(env.get("execution_policy", {}) or {})
    deferred_runner_kinds = policy.get("deferred_runner_kinds", [])
    return {
        "execution_stage": str(env.get("execution_stage", policy.get("stage", ""))).strip().lower(),
        "defer_stability_cases": bool(policy.get("defer_stability_cases", False)),
        "deferred_runner_kinds": set(deferred_runner_kinds or DEFERRED_STABILITY_RUNNER_KINDS),
    }


def should_defer_stability_case(rule: Optional[dict], env: Optional[dict] = None) -> bool:
    rule = rule or {}
    runner_kind = str(rule.get("runner_kind", "")).strip()
    if not runner_kind:
        return False
    policy = stability_execution_policy(env)
    return policy["defer_stability_cases"] and runner_kind in policy["deferred_runner_kinds"]



def classify_doc_case(case: DocCase, env: Optional[dict] = None) -> dict:
    env = env or {}
    text = case_text(case)
    current_wakeup = env.get("current_wakeup_word", "")
    wifi_state = str(env.get("wifi_state", "offline")).lower()

    applicability = classify_model_applicability(case, env=env)
    if applicability:
        return applicability

    if case.case_id in SUPPORTED_DOC_CASES:
        rule = SUPPORTED_DOC_CASES[case.case_id]
        if should_defer_stability_case(rule, env):
            return {
                "status": "skip",
                "reason": "当前处于调试阶段，稳定性/压测用例暂不纳入执行；待全链路走通后再恢复。",
                "runner_kind": rule["runner_kind"],
            }
        return {
            "status": "auto_executable_now",
            "reason": rule["notes"],
            "runner_kind": rule["runner_kind"],
        }

    if case.case_id in UNSUPPORTED_REASON_OVERRIDES:
        return {
            "status": "skip",
            "reason": UNSUPPORTED_REASON_OVERRIDES[case.case_id],
            "runner_kind": "",
        }

    if any(keyword.lower() in text.lower() for keyword in OTA_KEYWORDS):
        return {"status": "skip", "reason": "涉及 OTA/升级流程，当前阶段不自动执行高风险升级用例。", "runner_kind": ""}

    if case.level3 == MODE_ONLINE or any(token.kind.startswith("online_") for token in case.tokens):
        if wifi_state != "online":
            return {"status": "skip", "reason": f"依赖在线链路，当前设备状态为 wifi {wifi_state or 'offline'}。", "runner_kind": ""}

    if any(keyword in text for keyword in APP_KEYWORDS):
        return {"status": "skip", "reason": "依赖手机 APP 配置或联动，当前未自动接管 APP 侧动作。", "runner_kind": ""}

    if any(keyword in text for keyword in POWER_KEYWORDS):
        return {"status": "skip", "reason": "依赖断电/上电/掉电动作，当前持续日志会话下不自动执行电源循环。", "runner_kind": ""}

    if any(keyword in text for keyword in REMOTE_KEYWORDS):
        return {"status": "skip", "reason": "依赖遥控器或面板按键动作，当前未接入自动化外设。", "runner_kind": ""}

    if any(keyword in text for keyword in NETWORK_KEYWORDS):
        return {"status": "skip", "reason": "依赖联网/云端/路由器/网络切换，当前阶段仅推进离线链路。", "runner_kind": ""}

    if any(keyword in text for keyword in MANUAL_KEYWORDS) and not case.tokens:
        return {"status": "skip", "reason": "用例主要依赖人工听音/观察，暂不具备完全自动判定条件。", "runner_kind": ""}

    wake_tokens = [token.value for token in case.tokens if token.kind == "Wakeup" and token.channel == "talk"]
    if wake_tokens and current_wakeup and any(token != current_wakeup for token in wake_tokens):
        return {"status": "skip", "reason": f"当前唤醒词为 {current_wakeup}，用例要求的唤醒词与当前设备不一致。", "runner_kind": ""}

    if any(token.kind == "Action" and token.channel == "shell" and token.value not in SAFE_SHELL_COMMANDS for token in case.tokens):
        return {"status": "skip", "reason": "包含未验证或高风险串口命令，当前先跳过。", "runner_kind": ""}

    if case.tokens:
        return {"status": "skip", "reason": "当前未为该 doc 用例建立稳定自动判定规则，先保守跳过。", "runner_kind": ""}

    return {"status": "skip", "reason": "无可执行 token，且主要依赖人工或外部系统。", "runner_kind": ""}



def parse_tone_catalog(header_path: Optional[Path] = None) -> Dict[int, str]:
    header_path = header_path or (workspace_root() / "doc" / "reference" / "tone.h")
    mapping: Dict[int, str] = {}
    pattern = re.compile(r"TONE_ID_(\d+)\s*=\s*(\d+),//\s*(.+)")
    with header_path.open("r", encoding="utf-8", errors="replace") as handle:
        for raw in handle:
            match = pattern.search(raw)
            if not match:
                continue
            mapping[int(match.group(2))] = match.group(3).strip()
    return mapping



def load_env() -> dict:
    env_path = workspace_root() / "config" / "polaris_env.json"
    env = json.loads(env_path.read_text(encoding="utf-8"))
    model_name, model_source = infer_device_model(env)
    if model_name:
        env["current_device_model"] = model_name
        env["current_device_model_source"] = model_source
        env["device_capability_tags"] = build_device_capability_tags(model_name)
    return env
