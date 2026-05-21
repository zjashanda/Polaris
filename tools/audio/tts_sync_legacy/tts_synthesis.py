# coding=UTF-8
from ws4py.client.threadedclient import WebSocketClient
import sys
import hashlib
import base64
import glob
import websockets
import asyncio
import datetime
import openpyxl
import win32com
from hashlib import sha1
import json,jsonpath, time, threading, os

base_url = "ws://wsapi.xfyun.cn/v1/aiui"
#base_url = "ws://tts-api.xfyun.cn/v2/tts"

# 账号
# 海尔
app_id = "5af3aa4f"
api_key = "fe85d97976354eeeaf3d0122fb44ba2b"
# 美的
#app_id = "5e017c12"
#api_key = "b1c9db5797d7bd16567bb2e6a34cf075"

# 聆思
# app_id = "3899071b"
# api_key = "96e150958f1f5df3ba06bfefb5952134"

# 聆思2
# app_id = "6396ed00"
# api_key = "b54737b99bfc33825c17a80708eba936"

# 待合成文本txt路径
tts_path = r'.\tts.txt'
# tts_path = r"E:\Cloud_recognition\tools\海尔测试工具\tts\tts.txt"
# 合成音频存放路径
audio_path = r"tts\\"
# 发音人
    # 山东 x2_xiaodong_haier
    # 粤语 x2_xiaoyue_haier
    # 四川 x2_haier_yezi_sc
    # 上海 x2_ShCn_ziling
    # 普通话 x2_yezi_haier x_xiaoshi_haier x2_chongchong_haier 
    # 情感发音人 x2_xiaoye_haier
# ent_vcn = 'x2_xiaoshi_cts'
# ent_vcn = 'x_xiaoshi_haier'
ent_vcn = 'x4_yezi'
# ent_vcn = "x4_qianqian"
# ent_vcn = "x2_yezi"
# ent_vcn = "x4_yuexiaoni_assist"
# ent_vcn = "x_dangdang"
# ent_vcn = 'x2_chongchong'
# mp3 or pcm
audio_format = 'pcm'
# audio_format = 'mp3'
# 如何命名:text or sid
name_mode = 'text'

# 是否需要生成excel、excel路径
tts_excel = 0
workbook_path = audio_path + "tts.xlsx"

#结束标记
end_tag = "--end--"

def transform(xlsx_path):
    nameList = xlsx_path.split('\\')   #将path各部分分开
    parent_path = '\\'.join(nameList[0:-1])
    xlsx_name = nameList[-1].split('.')[0]
    excel=win32com.client.DispatchEx('excel.application')
    pro=excel.Workbooks.Open(xlsx_path)   #打开要转换的excel
    pro.SaveAs(parent_path + xlsx_name + ".xls", FileFormat=56)  # 另存为xls格式
    pro.Close()
    excel.Application.Quit()
    if os.path.exists(parent_path + xlsx_name + ".xls"):
        os.remove(xlsx_path)

async def ttsRequest(websocket, line):
        #打开文件读取内容
        # newline = line+"[p1000]"+line+"[p1000]"+line+"[p1000]"+line+"[p1000]"+line+"[p1000]"
        # await websocket.send(newline)
        await websocket.send(line)

        
async def ttsResponse(websocket, line):
    response = await websocket.recv()
    
    # file_name = str(line_num).zfill(3)+'_'+line.strip()[8] + "." + audio_format
    # file_name = str(line_num-1).zfill(3) + ".pcm"
    # 纯数字：file_name = str(line_num) + "." + audio_format
    # 纯文本：file_name = line.replace('\n','').replace('\r','')
    # file_name = line.strip() + "." + audio_format
    file_name = str(line_num-1).zfill(3)+'_'+line.strip() + "." + audio_format

    file_name = file_name.replace('，',',')
    file_name = file_name.replace(' ','')

    start_char = '['
    end_char = ']'
    if file_name.find(start_char) != -1:
        substring = file_name[file_name.index(start_char):file_name.index(end_char)+1]
        # print(substring)
        file_name = file_name.replace(substring,"")
    else:
        file_name = file_name.replace('[','')
        file_name = file_name.replace(']','')

    fp = open(audio_path + file_name, 'wb') 
    while response is not None and response !='':
        
        response = await websocket.recv()
        # print(response)
        jsonRslt = json.loads(response)
        if(name_mode=='sid'):
            new_name=jsonRslt["sid"] + "."+audio_format
        # print(jsonRslt)
        is_finish = jsonpath.jsonpath(jsonRslt, '$..data.is_finish')

        content = jsonpath.jsonpath(jsonRslt, '$.data.content')
        #print(content[0])
        base64Content = base64.b64decode(content[0])

        fp.write(base64Content)
        fp.flush()           
        if is_finish[0] == True:
            print("文件 " + file_name + " 合成完成。")
            fp.close()
            if(name_mode=='sid'):
                os.rename(os.path.join(audio_path + file_name),os.path.join(audio_path + new_name))

            if tts_excel:
                worksheet.cell(row=line_num,column=1).value = new_name
                worksheet.cell(row=line_num,column=2).value = line
            return

async def processTask(ttsContent):
    curTime = int(time.time())
    #\"ent\":\"xtts\",x2_yezi_sc，x2_xiaodong_haier

    #非情感发音人
    if(ent_vcn != 'x2_xiaoye_haier'):
        #pcm
        if(audio_format =='pcm'):
            if(ent_vcn == 'x2_yezi_haier' or ent_vcn == 'x2_yezi_sc_haier'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"45\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
            elif(ent_vcn == 'x4_qianqian'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"        
            elif(ent_vcn == 'x4_yezi'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"        
            elif(ent_vcn == 'x4_yuexiaoni_assist'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"        
            elif(ent_vcn == 'x_dangdang'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"        
            else:
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"60\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"        
        #mp3
        elif(audio_format =='mp3'):
            if(ent_vcn == 'x2_yezi_haier' or ent_vcn == 'x2_yezi_sc_haier'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"45\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
            elif(ent_vcn == 'x4_qianqian'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"45\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
            elif(ent_vcn == 'x4_yezi'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"45\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
            elif(ent_vcn == 'x4_yuexiaoni_assist'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
            elif(ent_vcn == 'x_dangdang'):
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
            else:
                param = "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"" + ent_vcn + "\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
        else:
            print('invalid audio format!')
    #情感发音人
    else:
        if(audio_format =='pcm'):
            param =  "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"x2_xiaoye_haier\",\"aue\":\"raw\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"attach_params\":\"{\\\"tts_params\\\":\\\"tts_emotion=1,tts_emotion_scale=10\\\"}\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
        elif(audio_format =='mp3'):
            param =  "{\"auth_id\":\"894c985bf8b1111c6728db79d3479aeg\",\"data_type\":\"text\",\"speed\":\"50\",\"pitch\":\"50\",\"volume\":\"100\",\"ent\":\"xtts\",\"vcn\":\"x2_xiaoye_haier\",\"aue\":\"lame\",\"sfl\":\"1\",\"scene\":\"IFLYTEK.tts\",\"sample_rate\":\"16000\",\"vad_info\":\"end\",\"attach_params\":\"{\\\"tts_params\\\":\\\"tts_emotion=1,tts_emotion_scale=10\\\"}\",\"ver_type\":\"monitor\",\"result_level\":\"plain\"}"
        else:
            print('invalid audio format!')
    #print('param:\n',param)
    #print(ttsContent)
    
    param = param.encode(encoding="utf-8")
    paramBase64 = base64.b64encode(param).decode()
    checkSumPre = api_key + str(curTime) + paramBase64
    checksum = hashlib.md5(checkSumPre.encode("utf-8")).hexdigest()
    connParam = "?appid="+app_id+"&checksum="+checksum+"&param="+paramBase64+"&curtime="+str(curTime)+"&signtype=md5"

    async with websockets.connect(base_url + connParam, origin="*",close_timeout=1000) as websocket:

        try:
            senAudioTask1 = asyncio.create_task(ttsRequest(websocket, ttsContent))
            recvRsltTask1 = asyncio.create_task(ttsResponse(websocket, ttsContent))

            await senAudioTask1
            await recvRsltTask1
        except Exception as e:
            print(e)
        print('send finish')
if __name__ == '__main__':

    ttsFile = open(tts_path, 'r+', encoding='utf-8')
    if tts_excel:
        if os.path.exists(workbook_path):
            workbook = openpyxl.load_workbook(workbook_path)
            worksheet = workbook.active
            line_num=worksheet.max_row+1
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            line_num=worksheet.max_row+1
            worksheet.cell(row=1,column=1).value = '【name】'
            worksheet.cell(row=1,column=2).value = '【期望结果】'
    else:
        line_num=1
    print('line_num:',line_num)
    
    for line in ttsFile:
        print(line)
        asyncio.run(processTask(line))
        line_num=line_num+1


    ttsFile.close()
    if tts_excel:
        workbook.save(workbook_path)
        workbook.close()
        transform(workbook_path)
    
